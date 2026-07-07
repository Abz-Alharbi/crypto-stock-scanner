import base64
import io
from pathlib import Path

import pytest
from openpyxl import load_workbook
from PIL import Image

from backend.errors import ApiError
from backend.services import pattern_detection


def _bearer(token):
    return {"Authorization": f"Bearer {token}"}


def _sample_png_base64():
    buffer = io.BytesIO()
    Image.new("RGB", (32, 32), color=(255, 255, 255)).save(buffer, format="PNG")
    return base64.b64encode(buffer.getvalue()).decode("ascii")


def test_pattern_detect_endpoint_accepts_base64_image_and_logs_detection(client, app, monkeypatch, tmp_path):
    app.config["PATTERN_LOG_ROOT"] = str(tmp_path / "pattern_detections")
    monkeypatch.setattr(pattern_detection, "get_redis_client", lambda: None)
    monkeypatch.setattr(
        pattern_detection.yoloService,
        "detect_patterns",
        lambda mat: [{"label": "Double Top", "confidence": 0.88, "bbox": [2, 3, 20, 21]}],
    )
    monkeypatch.setattr(pattern_detection.scans, "get_bars", lambda symbol, timeframe: [{"c": 1}] * 60)
    monkeypatch.setattr(
        pattern_detection.scans.ta,
        "full_analysis",
        lambda _bars: {
            "patterns": {
                "candlestick": [],
                "chart": [{"pattern": "double_top", "type": "bearish", "strength": "strong"}],
            }
        },
    )

    register_response = client.post(
        "/api/auth/register",
        json={"username": "pattern_user", "email": "pattern@example.test", "password": "SecurePass123"},
    )
    token = register_response.get_json()["token"]

    response = client.post(
        "/api/patterns/detect",
        json={"image": _sample_png_base64(), "symbol": "AAPL", "timeframe": "1D"},
        headers=_bearer(token),
    )

    assert response.status_code == 200
    payload = response.get_json()
    assert payload["signal_priority"] == 1
    assert payload["pattern"] == "Double Top"
    assert payload["source_badge"] == "YOLOv8 + TA-Lib"
    assert payload["talib_conflict"] is False
    assert payload["screenshot_path"].endswith(".jpg")

    screenshot_parts = Path(payload["screenshot_path"]).parts
    user_id = screenshot_parts[screenshot_parts.index("pattern_detections") + 1]
    assert Path(payload["screenshot_path"]).exists()
    log_files = list((tmp_path / "pattern_detections" / user_id).glob("*.xlsx"))
    assert len(log_files) == 1
    worksheet = load_workbook(log_files[0]).active
    rows = list(worksheet.iter_rows(values_only=True))
    assert rows[0] == tuple(pattern_detection.PATTERN_LOG_HEADERS)
    assert rows[1][1:7] == ("AAPL", "1D", "Double Top", 0.88, "YOLOv8 + TA-Lib", False)


def test_pattern_rate_limit_rejects_eleventh_request(monkeypatch):
    class FakeRedis:
        def __init__(self):
            self.count = 0

        def incr(self, _key):
            self.count += 1
            return self.count

        def expire(self, _key, _seconds):
            return True

        def ttl(self, _key):
            return 42

    redis = FakeRedis()
    monkeypatch.setattr(pattern_detection, "get_redis_client", lambda: redis)

    for _index in range(10):
        pattern_detection.ensure_pattern_rate_limit(7)

    with pytest.raises(ApiError) as exc:
        pattern_detection.ensure_pattern_rate_limit(7)

    assert exc.value.status_code == 429
    assert exc.value.details["limit"] == 10
