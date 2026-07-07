import base64
import binascii
import io
import logging
from datetime import datetime
from pathlib import Path

import numpy as np
from flask import current_app
from redis.exceptions import RedisError

from backend.errors import ApiError
from backend.services import scans
from backend.services.patternDetection import signalResolver, yoloService
from backend.services.redis_store import get_redis_client
from backend.symbols import canonicalize_symbol

logger = logging.getLogger(__name__)

PATTERN_RATE_LIMIT = 10
PATTERN_RATE_WINDOW_SECONDS = 60
PROJECT_ROOT = Path(__file__).resolve().parents[2]
PATTERN_LOG_ROOT = Path("logs") / "pattern_detections"
PATTERN_LOG_HEADERS = [
    "timestamp",
    "symbol",
    "timeframe",
    "pattern",
    "confidence",
    "source_badge",
    "talib_conflict",
    "screenshot_path",
]


def _import_cv2():
    try:
        import cv2
    except Exception as exc:
        details = {
            "exception_type": type(exc).__name__,
            "exception_message": str(exc),
        }
        logger.exception(
            "opencv_import_failed",
            extra=details,
        )
        raise ApiError("OpenCV is not installed", 503, "opencv_unavailable", details) from exc
    return cv2


def detect_pattern_for_user(user, data):
    mat = decode_base64_image(data.image)
    try:
        yolo_results = yoloService.detect_patterns(mat)
    except RuntimeError as exc:
        details = getattr(yoloService.get_yolo_service(), "last_error", None)
        message = str(exc)
        if isinstance(details, dict) and details.get("exception_message"):
            message = f"{message} {details['exception_message']}"
        raise ApiError(message, 503, "yolo_unavailable", details) from exc
    talib_patterns = []
    symbol = data.symbol
    timeframe = data.timeframe

    if symbol and timeframe:
        canonical = canonicalize_symbol(symbol)
        symbol = canonical.provider_symbol
        talib_patterns = get_talib_patterns(symbol, timeframe)

    signal = signalResolver.resolve(yolo_results, talib_patterns)
    timestamp = datetime.utcnow()
    screenshot_path = save_annotated_screenshot(user.id, timestamp, mat, signal["bounding_boxes"])
    append_detection_log(user.id, timestamp, symbol, timeframe, signal, screenshot_path)

    return {
        **signal,
        "symbol": symbol,
        "timeframe": timeframe,
        "screenshot_path": screenshot_path,
        "yolo_results": yolo_results,
        "talib_patterns": talib_patterns,
    }


def decode_base64_image(encoded_image):
    payload = encoded_image.split(",", 1)[1] if encoded_image.startswith("data:") else encoded_image
    try:
        image_bytes = base64.b64decode(payload, validate=True)
    except (ValueError, binascii.Error) as exc:
        raise ApiError("Invalid base64 image", 400, "invalid_image") from exc

    try:
        from PIL import Image, UnidentifiedImageError
    except Exception as exc:
        details = {
            "exception_type": type(exc).__name__,
            "exception_message": str(exc),
        }
        logger.exception("pillow_import_failed", extra=details)
        raise ApiError("Pillow is not installed", 503, "pillow_unavailable", details) from exc

    try:
        with Image.open(io.BytesIO(image_bytes)) as image:
            mat = np.asarray(image.convert("RGB")).copy()
    except (UnidentifiedImageError, OSError, ValueError):
        raise ApiError("Image could not be decoded", 400, "invalid_image")
    return mat


def get_talib_patterns(symbol, timeframe):
    bars = scans.get_bars(symbol, timeframe)
    analysis = scans.ta.full_analysis(bars)
    if not analysis:
        return []
    patterns = analysis.get("patterns") or {}
    return (patterns.get("candlestick") or []) + (patterns.get("chart") or [])


def save_annotated_screenshot(user_id, timestamp, image, bounding_boxes):
    try:
        from PIL import Image, ImageDraw
    except Exception as exc:
        details = {
            "exception_type": type(exc).__name__,
            "exception_message": str(exc),
        }
        logger.exception("pillow_import_failed", extra=details)
        raise ApiError("Pillow is not installed", 503, "pillow_unavailable", details) from exc

    root = _pattern_log_root()
    screenshot_dir = root / str(user_id) / "screenshots"
    screenshot_dir.mkdir(parents=True, exist_ok=True)
    screenshot_path = screenshot_dir / f"{_timestamp_slug(timestamp)}.jpg"

    array = np.asarray(image)
    if array.ndim == 2:
        annotated = Image.fromarray(array).convert("RGB")
    else:
        annotated = Image.fromarray(array.astype(np.uint8)).convert("RGB")
    draw = ImageDraw.Draw(annotated)
    for bbox in bounding_boxes or []:
        x1, y1, x2, y2 = [int(round(value)) for value in bbox]
        draw.rectangle((x1, y1, x2, y2), outline=(0, 255, 0), width=2)

    try:
        annotated.save(screenshot_path, format="JPEG", quality=90)
    except OSError:
        raise ApiError("Annotated screenshot could not be saved", 500, "screenshot_save_failed")
    return str(screenshot_path)


def append_detection_log(user_id, timestamp, symbol, timeframe, signal, screenshot_path):
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError as exc:
        raise ApiError("openpyxl is not installed", 503, "xlsx_unavailable") from exc

    log_dir = _pattern_log_root() / str(user_id)
    log_dir.mkdir(parents=True, exist_ok=True)
    log_path = log_dir / f"{timestamp.strftime('%Y-%m-%d')}.xlsx"

    if log_path.exists():
        workbook = load_workbook(log_path)
        worksheet = workbook.active
    else:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "detections"
        worksheet.append(PATTERN_LOG_HEADERS)

    worksheet.append(
        [
            timestamp.isoformat(),
            symbol,
            timeframe,
            signal.get("pattern"),
            signal.get("confidence"),
            signal.get("source_badge"),
            signal.get("talib_conflict"),
            screenshot_path,
        ]
    )
    workbook.save(log_path)
    return str(log_path)


def ensure_pattern_rate_limit(user_id):
    client = get_redis_client()
    if client is None:
        return

    key = f"rate_limit:patterns:{user_id}:{int(datetime.utcnow().timestamp() // PATTERN_RATE_WINDOW_SECONDS)}"
    try:
        count = client.incr(key)
        if count == 1:
            client.expire(key, PATTERN_RATE_WINDOW_SECONDS)
        if count > PATTERN_RATE_LIMIT:
            retry_after = client.ttl(key)
            raise ApiError(
                "Pattern detection rate limit exceeded",
                429,
                "rate_limited",
                {"limit": PATTERN_RATE_LIMIT, "window_seconds": PATTERN_RATE_WINDOW_SECONDS, "retry_after": retry_after},
            )
    except RedisError as exc:
        raise ApiError("Rate limit store is unavailable", 503, "rate_limit_unavailable") from exc


def _pattern_log_root():
    root = Path(current_app.config.get("PATTERN_LOG_ROOT", PATTERN_LOG_ROOT))
    return root if root.is_absolute() else PROJECT_ROOT / root


def _timestamp_slug(timestamp):
    return timestamp.strftime("%Y%m%dT%H%M%S%fZ")
